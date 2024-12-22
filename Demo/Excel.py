import openpyxl
book=openpyxl.load_workbook("C:\\Users\\Manpreet.Singh1\\Music\\Bhai.xlsx")
sheet=book.active
cell=sheet.cell(row=2,column=1)
'''print(cell.value)
#sheet.cell(row=3,column=1).value="Bubban"
print(sheet.cell(row=3,column=1).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['c2'].value)'''
dicti={}
for r in range(1, sheet.max_row+1):
    if sheet.cell(row=r,column=1).value=="There":
        for c in range(1,sheet.max_column+1):
            dicti[sheet.cell(row=1,column=c).value]=sheet.cell(row=r,column=c).value

print(dicti)
