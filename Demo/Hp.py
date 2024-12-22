from selenium import webdriver
import time
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
book=openpyxl.load_workbook("C:\\Users\\Manpreet.Singh1\\Music\\download.xlsx")
dict={}
driver=webdriver.Chrome()
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.implicitly_wait(5)
time.sleep(5)
fruitName='Apple'
driver.find_element(By.CSS_SELECTOR,"button[id='downloadButton']").click()
time.sleep(5)
driver.find_element(By.CSS_SELECTOR,"input[type='file']").send_keys("C:/Users/Manpreet.Singh1/Music/download.xlsx")
wait=WebDriverWait(driver,5)
loc=(By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(loc))
print(driver.find_element(*loc).text)
PriceColumn=driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
FruitPrice=driver.find_element(By.XPATH,("//div[text()='"+fruitName+"']/parent::div/parent::div/div[@id='cell-"+PriceColumn+"-undefined']/div")).text
print(FruitPrice)
sheet=book.active
for c in range(1,sheet.max_column+1):
    if sheet.cell(row=1,column=c).value=="price":
        dict["column"]=c

for r in range(1, sheet.max_row + 1):
    for co in range(1,sheet.max_column+1):
        if sheet.cell(row=r,column=co).value=="Apple":
            dict["row"]=r

AppleCell=sheet.cell(row=dict["row"],column=dict["column"])
print(AppleCell.value)
AppleCell.value=10000
book.save("C:/Users/Manpreet.Singh1/Music/download.xlsx")
